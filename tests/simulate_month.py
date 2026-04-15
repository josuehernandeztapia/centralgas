"""
Synthetic Month Simulation — Demo del cierre diario automatizado.

Simula 31 días de operación de la estación Nacozari (EDS 5) con
7 escenarios distintos que cubren todos los bloques de reconciliación.
Genera un Excel de salida replicando el formato NatGas de 38 columnas.

Escenarios:
  1. Día normal — todo cuadra (status OK)
  2. Diferencia centavos — redondeo típico ±$0.50 (status OK)
  3. Faltante TPV — tiras auditoras $80 arriba (status WARNING)
  4. Depósito bancario partido — 2 fechas, ingreso desfasado (status WARNING)
  5. Faltante efectivo — $600 de diferencia (status CRITICAL)
  6. SCADA offline — datos parciales (status PARTIAL)
  7. Sospecha fraude — $5,200 faltante bancario (status EMERGENCY)
"""

import sys
import random
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from app.models.transaction import MedioPago, SchemaVersion, TransactionNormalized, CST
from app.models.reconciliation import CloseStatus, ReconciliationThresholds
from app.services.reconciliation import run_daily_close, format_summary_message

random.seed(42)

# ============================================================
# Transaction generator — synthetic but realistic
# ============================================================

PLACAS = [f"AGS{i:04d}" for i in range(1, 201)]
PVPS = [Decimal("13.99"), Decimal("14.20"), Decimal("13.85")]

def generate_day_transactions(
    close_date: date,
    station_id: int = 3,
    num_cargas: int = 120,
    pct_efectivo: float = 0.84,
    pct_tarjeta: float = 0.10,
    pct_prepago: float = 0.04,
    pct_credito: float = 0.02,
) -> list[TransactionNormalized]:
    txns = []
    for i in range(num_cargas):
        r = random.random()
        if r < pct_efectivo:
            mp = MedioPago.EFECTIVO
        elif r < pct_efectivo + pct_tarjeta:
            mp = MedioPago.TARJETA_DEBITO
        elif r < pct_efectivo + pct_tarjeta + pct_prepago:
            mp = MedioPago.PREPAGO
        else:
            mp = MedioPago.CREDITO

        litros = Decimal(str(round(random.uniform(10, 50), 4)))
        pvp = random.choice(PVPS)
        total = (litros * pvp).quantize(Decimal("0.01"))
        recaudo = Decimal(str(round(random.uniform(0, 30), 2))) if random.random() < 0.15 else Decimal("0")
        placa = random.choice(PLACAS)
        hour = random.randint(5, 22)
        minute = random.randint(0, 59)
        local_dt = datetime(close_date.year, close_date.month, close_date.day, hour, minute, 0, tzinfo=CST)
        utc_dt = local_dt.astimezone(timezone.utc)
        ingreso_neto = (total / Decimal("1.16")).quantize(Decimal("0.01"))
        iva = total - ingreso_neto
        kg = (litros * Decimal("0.717")).quantize(Decimal("0.0001"))

        txns.append(TransactionNormalized(
            source_file="synthetic.csv", source_row=i + 1,
            schema_version=SchemaVersion.POST_2023,
            station_id=station_id, station_natgas="EDS Nacozari",
            plaza="AGUASCALIENTES",
            timestamp_utc=utc_dt, timestamp_local=local_dt,
            placa=placa, litros=litros, pvp=pvp,
            total_mxn=total, recaudo_pagado=recaudo,
            medio_pago=mp, kg=kg, nm3=litros,
            ingreso_neto=ingreso_neto, iva=iva,
        ))
    return txns


# ============================================================
# Scenario definitions
# ============================================================

def _tpv_from_txns(txns):
    """Sum TPV total from transactions."""
    return Decimal(str(round(sum(
        float(t.total_mxn) for t in txns
        if t.medio_pago in (MedioPago.TARJETA_DEBITO, MedioPago.TARJETA_CREDITO)
    ), 2)))

def scenario_normal(close_date, txns, gasup_ventas):
    """Escenario 1: Todo cuadra perfectamente."""
    efectivo = float(gasup_ventas)
    return dict(
        compusafe_efectivo=Decimal(str(round(efectivo - 45, 0))),
        compusafe_corte=Decimal("500000"),
        compusafe_retiro=Decimal(str(round(efectivo - 45, 0))),
        banco_monto=Decimal(str(round(efectivo - 45, 0))),
        banco_ingreso=Decimal(str(round(efectivo - 45, 0))),
        fecha_conciliado=close_date + timedelta(days=3),
        fecha_deposito=str(close_date),
        scada_nm3=None,
        tpv_tiras=_tpv_from_txns(txns),
    )

def scenario_centavos(close_date, txns, gasup_ventas):
    """Escenario 2: Redondeo de centavos típico (±$0.55)."""
    efectivo = float(gasup_ventas)
    offset = round(random.uniform(-0.55, 0.55), 2)
    deposited = round(efectivo - 43 + offset, 0)
    return dict(
        compusafe_efectivo=Decimal(str(deposited)),
        compusafe_corte=Decimal("510000"),
        compusafe_retiro=Decimal(str(deposited)),
        banco_monto=Decimal(str(deposited)),
        banco_ingreso=Decimal(str(deposited)),
        fecha_conciliado=close_date + timedelta(days=2),
        fecha_deposito=str(close_date),
        scada_nm3=None,
        tpv_tiras=_tpv_from_txns(txns),
    )

def scenario_tpv_warning(close_date, txns, gasup_ventas):
    """Escenario 3: Tiras auditoras TPV $80 arriba — WARNING."""
    efectivo = float(gasup_ventas)
    deposited = round(efectivo - 35, 0)
    # Calculamos el TPV total de las transacciones para generar las tiras con offset
    tpv_txns = [t for t in txns if t.medio_pago in (MedioPago.TARJETA_DEBITO, MedioPago.TARJETA_CREDITO)]
    tpv_total = sum(float(t.total_mxn) for t in tpv_txns)
    return dict(
        compusafe_efectivo=Decimal(str(deposited)),
        compusafe_corte=Decimal("520000"),
        compusafe_retiro=Decimal(str(deposited)),
        banco_monto=Decimal(str(deposited)),
        banco_ingreso=Decimal(str(deposited)),
        fecha_conciliado=close_date + timedelta(days=2),
        fecha_deposito=str(close_date),
        scada_nm3=None,
        tpv_tiras=Decimal(str(round(tpv_total + 80, 2))),
    )

def scenario_split_deposit(close_date, txns, gasup_ventas):
    """Escenario 4: Depósito bancario partido en 2 fechas — ingreso desfasado."""
    efectivo = float(gasup_ventas)
    deposited = round(efectivo - 50, 0)
    half = round(deposited / 2, 0)
    return dict(
        compusafe_efectivo=Decimal(str(deposited)),
        compusafe_corte=Decimal("530000"),
        compusafe_retiro=Decimal(str(deposited)),
        banco_monto=Decimal(str(deposited)),
        banco_ingreso=Decimal(str(round(half + 300, 0))),  # Only half arrived + some from prev day
        fecha_conciliado=close_date + timedelta(days=5),
        fecha_deposito=f"{close_date} & {close_date + timedelta(days=2)}",
        scada_nm3=None,
        tpv_tiras=_tpv_from_txns(txns),
    )

def scenario_cash_short(close_date, txns, gasup_ventas):
    """Escenario 5: Faltante de efectivo $600 — CRITICAL."""
    efectivo = float(gasup_ventas)
    deposited = round(efectivo - 40 - 600, 0)  # $600 short
    return dict(
        compusafe_efectivo=Decimal(str(deposited)),
        compusafe_corte=Decimal("540000"),
        compusafe_retiro=Decimal(str(deposited)),
        banco_monto=Decimal(str(deposited)),
        banco_ingreso=Decimal(str(deposited)),
        fecha_conciliado=close_date + timedelta(days=2),
        fecha_deposito=str(close_date),
        scada_nm3=None,
        tpv_tiras=_tpv_from_txns(txns),
    )

def scenario_scada_offline(close_date, txns, gasup_ventas):
    """Escenario 6: SCADA offline — datos parciales."""
    efectivo = float(gasup_ventas)
    deposited = round(efectivo - 30, 0)
    return dict(
        compusafe_efectivo=Decimal(str(deposited)),
        compusafe_corte=Decimal("550000"),
        compusafe_retiro=Decimal(str(deposited)),
        banco_monto=Decimal(str(deposited)),
        banco_ingreso=Decimal(str(deposited)),
        fecha_conciliado=close_date + timedelta(days=2),
        fecha_deposito=str(close_date),
        scada_nm3=Decimal("0"),  # SCADA reported 0 = offline
        tpv_tiras=_tpv_from_txns(txns),
    )

def scenario_fraud(close_date, txns, gasup_ventas):
    """Escenario 7: Sospecha fraude — $5,200 faltante bancario — EMERGENCY."""
    efectivo = float(gasup_ventas)
    deposited = round(efectivo - 35, 0)
    return dict(
        compusafe_efectivo=Decimal(str(deposited)),
        compusafe_corte=Decimal("560000"),
        compusafe_retiro=Decimal(str(deposited)),
        banco_monto=Decimal(str(round(deposited - 5200, 0))),  # $5,200 missing from bank
        banco_ingreso=Decimal(str(round(deposited - 5200, 0))),
        fecha_conciliado=close_date + timedelta(days=1),
        fecha_deposito=str(close_date),
        scada_nm3=None,
        tpv_tiras=_tpv_from_txns(txns),
    )


SCENARIOS = [
    ("Normal", scenario_normal),
    ("Centavos", scenario_centavos),
    ("TPV Warning", scenario_tpv_warning),
    ("Split Deposit", scenario_split_deposit),
    ("Cash Short", scenario_cash_short),
    ("SCADA Offline", scenario_scada_offline),
    ("Fraud", scenario_fraud),
]


# ============================================================
# Run simulation
# ============================================================

def simulate_month(year=2026, month=1, station_id=3):
    """Simulate a full month of daily closes."""
    from calendar import monthrange
    num_days = monthrange(year, month)[1]

    results = []
    all_messages = []

    print(f"\n{'='*70}")
    print(f"  SIMULACIÓN MES COMPLETO — Enero {year} — EDS Nacozari")
    print(f"{'='*70}\n")

    for day in range(1, num_days + 1):
        close_date = date(year, month, day)
        scenario_name, scenario_fn = SCENARIOS[(day - 1) % len(SCENARIOS)]

        # Generate transactions (vary volume by day of week)
        dow = close_date.weekday()
        if dow < 5:
            num_cargas = random.randint(100, 150)
        else:
            num_cargas = random.randint(60, 90)

        txns = generate_day_transactions(close_date, station_id, num_cargas)

        # Calculate gasup ventas for scenario setup
        from app.services.reconciliation import build_gasup_block
        gasup_block = build_gasup_block(txns, close_date, station_id)
        gasup_ventas = gasup_block.ventas_mas_recaudos

        # Get scenario-specific Odoo/SCADA/manual values
        kwargs = scenario_fn(close_date, txns, gasup_ventas)

        # Run daily close
        result = run_daily_close(
            station_id=station_id,
            close_date=close_date,
            transactions=txns,
            **kwargs,
        )
        results.append((scenario_name, result))

        # Status emoji
        status_emoji = {
            CloseStatus.OK: "✅",
            CloseStatus.WARNING: "⚠️ ",
            CloseStatus.CRITICAL: "🚨",
            CloseStatus.EMERGENCY: "🔴",
            CloseStatus.PARTIAL: "⏳",
        }.get(result.status, "❓")

        # Print summary line
        checks_summary = ", ".join(
            f"{c.check_id.replace('CHK_','')}: {c.result.value}"
            for c in result.checks
            if c.result.value != "PASS"
        )
        if not checks_summary:
            checks_summary = "all pass"

        print(
            f"  {close_date.strftime('%d/%m')} {status_emoji} {result.status.value:10s} "
            f"| {scenario_name:14s} | {result.gasup.num_cargas:3d} cargas "
            f"| {result.gasup.total_litros:>8,.1f} lt "
            f"| ${result.gasup.total_mxn:>10,.2f} "
            f"| {checks_summary}"
        )

    # Summary stats
    print(f"\n{'─'*70}")
    status_counts = {}
    for _, r in results:
        status_counts[r.status.value] = status_counts.get(r.status.value, 0) + 1

    total_litros = sum(r.gasup.total_litros for _, r in results)
    total_mxn = sum(r.gasup.total_mxn for _, r in results)
    total_cargas = sum(r.gasup.num_cargas for _, r in results)
    dias_cerrados = sum(1 for _, r in results if r.dia_cerrado)

    print(f"\n  RESUMEN DEL MES:")
    print(f"  Total cargas:  {total_cargas:,}")
    print(f"  Total litros:  {total_litros:,.2f}")
    print(f"  Total ventas:  ${total_mxn:,.2f} MXN")
    print(f"  Días cerrados: {dias_cerrados}/{len(results)}")
    print(f"\n  Por status:")
    for status, count in sorted(status_counts.items()):
        pct = count / len(results) * 100
        print(f"    {status:12s}: {count:2d} días ({pct:.0f}%)")

    # Print one WhatsApp message sample
    print(f"\n{'─'*70}")
    print(f"  EJEMPLO MENSAJE WHATSAPP (Día 1 — Normal):\n")
    for line in results[0][1].whatsapp_message.split("\n"):
        print(f"  {line}")

    # Print a critical day message
    critical_days = [(name, r) for name, r in results if r.status == CloseStatus.CRITICAL]
    if critical_days:
        print(f"\n{'─'*70}")
        print(f"  EJEMPLO MENSAJE WHATSAPP (Día crítico — {critical_days[0][0]}):\n")
        for line in critical_days[0][1].whatsapp_message.split("\n"):
            print(f"  {line}")

    return results


# ============================================================
# Generate NatGas-format Excel output
# ============================================================

def generate_natgas_excel(results, output_path):
    """Generate Excel file replicating NatGas 38-column format."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ENERO 2026"

    # Styles
    header_font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    header_fill_blue = PatternFill("solid", fgColor="2F5496")
    header_fill_green = PatternFill("solid", fgColor="548235")
    header_fill_orange = PatternFill("solid", fgColor="BF8F00")
    header_fill_purple = PatternFill("solid", fgColor="7030A0")
    header_fill_red = PatternFill("solid", fgColor="C00000")
    header_fill_teal = PatternFill("solid", fgColor="2E75B6")
    header_fill_dark = PatternFill("solid", fgColor="404040")

    data_font = Font(name="Arial", size=9)
    money_fmt = '#,##0.00'
    pct_fmt = '0.00%'
    date_fmt = 'DD/MM/YYYY'

    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # ---- Row 1: Block group headers ----
    block_headers = [
        (1, 1, ""),
        (2, 16, "INFORMACIÓN POS (GasUp Agent)"),
        (17, 19, "COMPUSAFE"),
        (20, 23, "RESUMEN EFECTIVO"),
        (24, 26, "PICOS ETV"),
        (27, 32, "BANCOS"),
        (33, 34, "TPV AUDIT"),
        (35, 36, "INGRESOS"),
        (37, 37, ""),
        (38, 38, "STATUS"),
    ]
    fills = [None, header_fill_blue, header_fill_green, header_fill_orange,
             header_fill_purple, header_fill_teal, header_fill_dark,
             header_fill_red, None, header_fill_blue]

    for (start, end, title), fill in zip(block_headers, fills):
        if title:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            cell = ws.cell(row=1, column=start, value=title)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            if fill:
                for c in range(start, end + 1):
                    ws.cell(row=1, column=c).fill = fill

    # ---- Row 2: Column headers (38 columns matching NatGas) ----
    col_headers = [
        "FECHA",                           # A (1)
        "LITROS CONTADO",                  # B (2)
        "$ PV",                            # C (3)
        "LT DIFF PRECIO",                 # D (4)
        "$ PV DIFF",                       # E (5)
        "LT CRÉDITO",                      # F (6)
        "LT PREPAGO",                      # G (7)
        "TOTAL LITROS",                    # H (8)
        "VENTAS CANAST.",                  # I (9)
        "Desc. Efectivo",                  # J (10)
        "BONOS EDS",                       # K (11)
        "TPV",                             # L (12)
        "EDENRED",                         # M (13)
        "VENTA EFVO REAL",                 # N (14)
        "RECAUDOS",                        # O (15)
        "VENTAS+RECAUDOS",                 # P (16)
        "CORTE TEMPORAL",                  # Q (17)
        "RETIRO",                          # R (18)
        "EFVO COMPUSAFE",                  # S (19)
        "PICOS REALES",                    # T (20)
        "TOTAL EFVO OASIS",               # U (21)
        "EFVO OBTENIDO",                   # V (22)
        "DIFF EFECTIVO",                   # W (23)
        "DIA VENTA",                       # X (24)
        "COMPROBANTE",                     # Y (25)
        "IMPORTE",                         # Z (26)
        "FECHA CONCILIADO",               # AA (27)
        "FECHA DEPÓSITO",                  # AB (28)
        "BANCOS $",                        # AC (29)
        "VENTAS+REC CHECK",               # AD (30)
        "PICOS NECES.",                    # AE (31)
        "DIFF VS EFECTIVO",               # AF (32)
        "TIRAS AUDITORAS",                # AG (33)
        "VS GASDATA",                      # AH (34)
        "BANCOS INGRESO",                 # AI (35)
        "DIFF INGRESOS",                   # AJ (36)
        "ESCENARIO",                       # AK (37)
        "DIA CERRADO",                     # AL (38)
    ]

    for i, header in enumerate(col_headers, 1):
        cell = ws.cell(row=2, column=i, value=header)
        cell.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
        # Color by block
        if i <= 1: cell.fill = header_fill_dark
        elif i <= 16: cell.fill = header_fill_blue
        elif i <= 19: cell.fill = header_fill_green
        elif i <= 23: cell.fill = header_fill_orange
        elif i <= 26: cell.fill = header_fill_purple
        elif i <= 32: cell.fill = header_fill_teal
        elif i <= 34: cell.fill = header_fill_dark
        elif i <= 36: cell.fill = header_fill_red
        elif i == 37: cell.fill = PatternFill("solid", fgColor="808080")
        else: cell.fill = header_fill_blue

    # ---- Data rows (row 3+) ----
    for row_idx, (scenario_name, r) in enumerate(results, 3):
        g = r.gasup
        cs = r.cash_summary
        bk = r.bank
        tpv = r.tpv_audit
        inc = r.income
        etv = r.etv

        row_data = [
            r.close_date,                                    # A
            float(g.litros_contado),                         # B
            float(g.pvp_contado),                            # C
            float(g.litros_diff_precio),                     # D
            float(g.pvp_diff) if g.pvp_diff else 0,         # E
            float(g.litros_credito),                         # F
            float(g.litros_prepago),                         # G
            None,                                            # H = formula
            float(g.ventas_canastilla),                      # I
            float(g.descuentos_efectivo),                    # J
            float(g.bonos_eds),                              # K
            float(g.tpv_total),                              # L
            float(g.edenred),                                # M
            float(g.total_venta_efectivo),                   # N
            float(g.total_recaudos),                         # O
            None,                                            # P = formula
            float(r.compusafe.corte_temporal),               # Q
            float(r.compusafe.retiro),                       # R
            float(r.compusafe.efectivo_ingresado),           # S
            None,                                            # T = formula
            None,                                            # U = formula
            None,                                            # V = formula
            None,                                            # W = formula
            r.close_date,                                    # X
            etv.comprobante or "",                           # Y
            None,                                            # Z = formula
            bk.fecha_conciliado,                             # AA
            bk.fecha_deposito or "",                         # AB
            float(bk.banco_monto),                          # AC
            None,                                            # AD = formula
            None,                                            # AE = formula
            None,                                            # AF = formula
            float(tpv.tiras_auditoras),                     # AG
            None,                                            # AH = formula
            float(inc.bancos),                               # AI
            None,                                            # AJ = formula
            scenario_name,                                   # AK
            "SI" if r.dia_cerrado else "NO",                 # AL
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border

        # Excel formulas
        rr = row_idx
        ws.cell(row=rr, column=8).value = f'=B{rr}+D{rr}+F{rr}+G{rr}'        # H = B+D+F+G
        ws.cell(row=rr, column=16).value = f'=N{rr}+O{rr}'                     # P = N+O
        ws.cell(row=rr, column=20).value = f'=P{rr}-S{rr}'                     # T = P-S
        ws.cell(row=rr, column=21).value = f'=P{rr}'                           # U = P
        ws.cell(row=rr, column=22).value = f'=S{rr}+T{rr}'                     # V = S+T
        ws.cell(row=rr, column=23).value = f'=U{rr}-V{rr}'                     # W = U-V
        ws.cell(row=rr, column=26).value = f'=T{rr}'                           # Z = T
        ws.cell(row=rr, column=30).value = f'=P{rr}'                           # AD = P
        ws.cell(row=rr, column=31).value = f'=T{rr}'                           # AE = T
        ws.cell(row=rr, column=32).value = f'=AD{rr}-AC{rr}-AE{rr}'           # AF = AD-AC-AE
        ws.cell(row=rr, column=34).value = f'=AG{rr}-L{rr}'                   # AH = AG-L
        ws.cell(row=rr, column=36).value = f'=AI{rr}-(AC{rr}+Z{rr})'         # AJ = AI-(AC+Z)

        # Number formats
        for col in [2, 4, 6, 7, 8]:
            ws.cell(row=rr, column=col).number_format = '#,##0.00'
        for col in [3, 5]:
            ws.cell(row=rr, column=col).number_format = '$#,##0.00'
        for col in [9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 26, 29, 30, 31, 32, 33, 34, 35, 36]:
            ws.cell(row=rr, column=col).number_format = '$#,##0.00'

        # Conditional fill for status column (AL)
        al_cell = ws.cell(row=rr, column=38)
        if r.dia_cerrado:
            al_cell.fill = ok_fill
        elif r.status == CloseStatus.WARNING:
            al_cell.fill = warn_fill
        else:
            al_cell.fill = fail_fill

        # Highlight DIFF columns if non-zero warning/fail
        diff_cols = [23, 32, 34, 36]  # W, AF, AH, AJ
        diff_checks = ["CHK_CASH_BALANCE", "CHK_BANK_VS_CASH", "CHK_TPV_AUDIT", "CHK_INCOME"]
        for dc, check_id in zip(diff_cols, diff_checks):
            chk = next((c for c in r.checks if c.check_id == check_id), None)
            if chk and chk.result.value == "WARN":
                ws.cell(row=rr, column=dc).fill = warn_fill
            elif chk and chk.result.value == "FAIL":
                ws.cell(row=rr, column=dc).fill = fail_fill

    # ---- Totals row ----
    total_row = len(results) + 3
    ws.cell(row=total_row, column=1, value="TOTALES").font = Font(name="Arial", bold=True, size=9)
    for col in [2, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 19, 20, 21, 26, 29, 33, 35]:
        letter = get_column_letter(col)
        ws.cell(row=total_row, column=col).value = f'=SUM({letter}3:{letter}{total_row-1})'
        ws.cell(row=total_row, column=col).number_format = '$#,##0.00'
        ws.cell(row=total_row, column=col).font = Font(name="Arial", bold=True, size=9)

    # Count dias cerrados
    ws.cell(row=total_row, column=38).value = f'=COUNTIF(AL3:AL{total_row-1},"SI")&"/"&{total_row-3}'
    ws.cell(row=total_row, column=38).font = Font(name="Arial", bold=True, size=9)
    ws.cell(row=total_row, column=37, value="DÍAS CERRADOS:").font = Font(name="Arial", bold=True, size=9)

    # ---- Column widths ----
    widths = {
        'A': 12, 'B': 13, 'C': 8, 'D': 12, 'E': 8, 'F': 11, 'G': 11,
        'H': 13, 'I': 11, 'J': 12, 'K': 11, 'L': 11, 'M': 10,
        'N': 14, 'O': 11, 'P': 14, 'Q': 12, 'R': 12, 'S': 13,
        'T': 12, 'U': 14, 'V': 13, 'W': 13, 'X': 12, 'Y': 12,
        'Z': 11, 'AA': 13, 'AB': 16, 'AC': 12, 'AD': 14, 'AE': 12,
        'AF': 14, 'AG': 13, 'AH': 11, 'AI': 13, 'AJ': 13, 'AK': 13, 'AL': 12,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze first 2 rows + column A
    ws.freeze_panes = "B3"

    wb.save(output_path)
    return output_path


# ============================================================
# Main
# ============================================================

if __name__ == "__main__":
    results = simulate_month()

    output_path = Path("/sessions/cool-focused-pasteur/mnt/Downloads/central-gas-agent/output/Cierre_Enero_2026_Simulacion.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    generate_natgas_excel(results, output_path)
    print(f"\n  📊 Excel generado: {output_path}")
    print(f"  {len(results)} filas × 38 columnas — formato NatGas replicado")
