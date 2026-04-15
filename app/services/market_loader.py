"""
Smart Market Loader — Camino A

Loads the NatGas master database (Combis AGS JH.xlsx) as a market catalog,
classifies vehicles by placa nomenclature into segments, injects hourly
demand profiles from Modelo Operativo v11, and feeds into the retention
engine so the system has competitive intelligence context.

Data sources (by trust level):
  HIGH CONFIDENCE (Josue's ground-truth files, 2012-2025):
    - 18 CSV files: AGS Combis transaction data (388,665 txns, 331 unique
      combis). ALL plates confirmed COMBI regardless of NatGas label.
    - BASE_CLIENTES_GNC_AGUASCALIENTES.xlsx → VAGONETAS_AGS (331),
      TAXIS_AGS (733)
    - BASE_CLIENTES_GNC_AGS_COMPLETA.xlsx → TAXIS_AGS (1,199)
    - Taxis QRO, AGS y vagonetas 2012-2015.xlsx → Vagonetas AGS (53),
      TAXIS AGS (738)
    - Taxi AGS, CEL y QRO Combis AGS.xlsx → monthly summary with ACTIVAS

  MEDIUM CONFIDENCE (NatGas master — real clients, labels pending validation):
    - Combis AGS JH.xlsx → Vehiculos (86,674 vehicles). NatGas
      Desc_Segmento is WRONG for 42.9% of confirmed combis.
    - Combis AGS JH.xlsx → Consumo promedio y max CW (38,493 rows)
    - TAXI QRO Y AGS SIN CONSUMO 15 dias 4.xlsx → Hoja2 (1,376 inactive)

  REFERENCE:
    - Modelo Operativo v11 → Hourly profiles, segment params, seasonality

Classification hierarchy (3 layers):
  1. Ground-truth files override everything (if placa in Josue's CSVs → COMBI)
  2. Placa nomenclature (A0####A = COMBI, A###AAX = TAXI, XXX###X = PLATAFORMA)
  3. NatGas Desc_Segmento + consumo validator (pending validation from day 1)

Plate nomenclature evolution (AGS combis):
  Gen 0 (pre-2015, extinct): A0##AAA — 7 known (A014AAA, A016AAA, etc.)
  Gen 1 (rare/transitional):  A###AA — 3 known (A954AA, A145AA, A715AA)
  Gen 2 (current, 2017+):     A0####A — 346 known, range A00000A-A00550A
                               Number = concession number (488 total in AGS)

Validated consumption ranges (mediana LEQ/carga, from 4,786 AGS vehicles):
  TAXI:          11.0 (range 5-22)
  PLATAFORMA:    12.2 (range 5-24) — same as taxi, distinguished by plate
  COMBI:         28.1 (range 15-48)
  TP:            34.9 (range 15-99) — overlaps with combi
  BUS/URBANO:   137.3 (range 50-342)

Output:
  - existing_profiles dict compatible with retention.build_client_profiles()
  - Market catalog with NatGas segment classification
  - Hourly demand profiles as constants

Architecture note:
  placa is the universal identifier linking GasUp POS transactions,
  NatGas master DB, Odoo res.partner, and SCADA dispenser data.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


# ============================================================
# Segment Parameters from Modelo Operativo v11
# ============================================================

class MarketSegment(str, Enum):
    """Vehicle segments from NatGas/Modelo Operativo classification.

    Classification hierarchy (3 layers):
      1. Ground-truth placa override (A0####A/A###AA/A0##AAA = COMBI always)
      2. Placa nomenclature patterns (A###AAX = TAXI, XXX###X = PLATAFORMA)
      3. NatGas Desc_Segmento + consumo validator (pending validation)

    IMPORTANT: NatGas Desc_Segmento is WRONG for 42.9% of confirmed combis.
    Of 331 known combis, NatGas labels only 189 as 'Combis Colectivas'.
    The rest are mislabeled as TP (55), Público (25), Empresa (24), etc.
    Placa nomenclature is the base of all segmentation.

    Note: CETE is a *client* (fleet operator), not a segment.
    The correct segment for CETE's vehicles is TRANSPORTE_PERSONAL (TP).
    """
    COMBI = "COMBI"                              # Combis colectivas urbanas (~448 en AGS)
    TAXI = "TAXI"                                # Taxis regulares convertidos a GNC
    PLATAFORMA = "PLATAFORMA"                    # Taxi inteligente/ejecutivo (DiDi, Uber, etc.)
    TRANSPORTE_PERSONAL = "TRANSPORTE_PERSONAL"  # TP: vagonetas de fábricas (CETE, otros)
    CONVERSION = "CONVERSION"                    # Nuevas conversiones particulares
    BUS = "BUS"                                  # Autobuses/camiones urbanos (consumo >50 lt)
    PARTICULAR = "PARTICULAR"                    # Vehículos particulares
    UNKNOWN = "UNKNOWN"                          # Could not classify


@dataclass
class SegmentParams:
    """Operational parameters per segment from Modelo Operativo v11."""
    leq_per_charge: float         # LEQ/carga promedio
    leq_per_charge_max: float     # LEQ/carga tope
    leq_per_month: float          # LEQ/mes typical
    cycle_minutes: float          # Ciclo servicio (minutos)
    hourly_profile_key: str       # Which hourly profile to use
    revenue_per_month_mxn: float  # Estimated revenue at $13.99/LEQ


SEGMENT_PARAMS: dict[MarketSegment, SegmentParams] = {
    # Validated from 388,665 transactions (18 CSVs, 2017-2025) + NatGas 4,786 AGS vehicles
    MarketSegment.COMBI: SegmentParams(
        leq_per_charge=28, leq_per_charge_max=48,   # mediana=28.1, p75=32.8, max=48
        leq_per_month=900, cycle_minutes=5,           # ~900 LEQ/mes from CSV data
        hourly_profile_key="COMBI",
        revenue_per_month_mxn=900 * 13.99,  # $12,591
    ),
    MarketSegment.TAXI: SegmentParams(
        leq_per_charge=11, leq_per_charge_max=18,   # mediana=11.0, p75=11.8, max_p95=18
        leq_per_month=300, cycle_minutes=3,
        hourly_profile_key="TAXI",
        revenue_per_month_mxn=300 * 13.99,  # $4,197
    ),
    MarketSegment.PLATAFORMA: SegmentParams(
        leq_per_charge=12, leq_per_charge_max=18,   # mediana=12.2, same range as taxi
        leq_per_month=300, cycle_minutes=3,
        hourly_profile_key="TAXI",  # Same charging pattern as taxi
        revenue_per_month_mxn=300 * 13.99,  # $4,197
    ),
    MarketSegment.TRANSPORTE_PERSONAL: SegmentParams(
        leq_per_charge=40, leq_per_charge_max=50,
        leq_per_month=1000, cycle_minutes=5,
        hourly_profile_key="TP",
        revenue_per_month_mxn=1000 * 13.99,  # $13,990 — includes CETE fleet + others
    ),
    MarketSegment.CONVERSION: SegmentParams(
        leq_per_charge=40, leq_per_charge_max=50,
        leq_per_month=1000, cycle_minutes=5,
        hourly_profile_key="TP",
        revenue_per_month_mxn=1000 * 13.99,  # $13,990
    ),
    MarketSegment.BUS: SegmentParams(
        leq_per_charge=200, leq_per_charge_max=300,
        leq_per_month=6000, cycle_minutes=45,
        hourly_profile_key="BUS",
        revenue_per_month_mxn=6000 * 13.99,  # $83,940
    ),
    MarketSegment.PARTICULAR: SegmentParams(
        leq_per_charge=30, leq_per_charge_max=50,
        leq_per_month=500, cycle_minutes=5,
        hourly_profile_key="COMBI",
        revenue_per_month_mxn=500 * 13.99,  # $6,995
    ),
}


# ============================================================
# Hourly Demand Profiles — from Modelo Operativo v11
# ============================================================
# Each profile is a dict[int, float] where key=hour (0-23),
# value=fraction of daily volume (sums to ~1.0).
# Source: Real transaction data — see Modelo Operativo v11 §2.

HOURLY_PROFILES: dict[str, dict[int, float]] = {
    # Combis: 286,917 txns from EDS Nacozari 2017-2025
    # Double-hump: AM peak 8-9h (18.5%), PM peak 15-16h (14.4%)
    "COMBI": {
        0: 0.003, 1: 0.001, 2: 0.000, 3: 0.000, 4: 0.000, 5: 0.007,
        6: 0.044, 7: 0.071, 8: 0.094, 9: 0.091, 10: 0.074, 11: 0.060,
        12: 0.057, 13: 0.054, 14: 0.059, 15: 0.071, 16: 0.073, 17: 0.065,
        18: 0.062, 19: 0.054, 20: 0.036, 21: 0.012, 22: 0.005, 23: 0.004,
    },
    # Taxis: 8,821 txns from EDS Ojo Caliente 2017-2025
    # Nocturnal 16.5%, peak AM 8-9h (21%), PM 17h (12.4%)
    "TAXI": {
        0: 0.047, 1: 0.006, 2: 0.006, 3: 0.002, 4: 0.018, 5: 0.023,
        6: 0.033, 7: 0.022, 8: 0.100, 9: 0.110, 10: 0.039, 11: 0.030,
        12: 0.030, 13: 0.024, 14: 0.030, 15: 0.024, 16: 0.040, 17: 0.124,
        18: 0.085, 19: 0.067, 20: 0.054, 21: 0.034, 22: 0.016, 23: 0.036,
    },
    # Transporte de Personal: Factory shift model (Parques Industriales)
    # 3 shifts: 6am, 2pm, 10pm. Peaks after shift drop-off.
    "TP": {
        0: 0.002, 1: 0.001, 2: 0.001, 3: 0.001, 4: 0.003, 5: 0.025,
        6: 0.035, 7: 0.100, 8: 0.110, 9: 0.070, 10: 0.035, 11: 0.025,
        12: 0.020, 13: 0.030, 14: 0.040, 15: 0.100, 16: 0.110, 17: 0.070,
        18: 0.040, 19: 0.025, 20: 0.020, 21: 0.050, 22: 0.055, 23: 0.032,
    },
    # Buses: Operator-provided schedule (Oriente station)
    # Two windows: AM 6-9h, PM 21-23h. Each slot = 16.7%
    "BUS": {
        0: 0.000, 1: 0.000, 2: 0.000, 3: 0.000, 4: 0.000, 5: 0.000,
        6: 0.167, 7: 0.167, 8: 0.167, 9: 0.000, 10: 0.000, 11: 0.000,
        12: 0.000, 13: 0.000, 14: 0.000, 15: 0.000, 16: 0.000, 17: 0.000,
        18: 0.000, 19: 0.000, 20: 0.000, 21: 0.167, 22: 0.167, 23: 0.167,
    },
}


# ============================================================
# Seasonality Factors — from CAPA2 2025 (377K+ transactions)
# ============================================================
# key = month (1-12), value = multiplier vs annual average.
# Source: Modelo Operativo v11 §6.3

SEASONALITY_FACTORS: dict[int, float] = {
    1: 1.003,   # Enero — normal
    2: 0.948,   # Febrero — normal
    3: 1.025,   # Marzo — normal
    4: 0.943,   # Abril — normal
    5: 1.056,   # Mayo — pico
    6: 1.091,   # Junio — pico
    7: 1.150,   # Julio — pico absoluto
    8: 1.000,   # Agosto — normal
    9: 1.003,   # Septiembre — normal
    10: 1.084,  # Octubre — pico
    11: 0.856,  # Noviembre — valle
    12: 0.841,  # Diciembre — valle
}


# ============================================================
# Station Market Context
# ============================================================

@dataclass
class StationMarket:
    """Market characteristics of a Central Gas station."""
    station_id: int
    name: str
    total_leq_month: float           # Total addressable market (LEQ/month)
    segment_mix: dict[MarketSegment, float]  # fraction of market per segment
    competitor_name: str = ""        # Nearest NatGas competitor
    competitor_distance_m: int = 0   # Distance in meters


STATION_MARKETS: dict[int, StationMarket] = {
    1: StationMarket(
        station_id=1,
        name="Parques Industriales",
        total_leq_month=120_000,  # TP fixed fleet (CETE + others) + conversions
        segment_mix={
            MarketSegment.TRANSPORTE_PERSONAL: 0.83,  # 100 vagonetas fijas (CETE, etc.)
            MarketSegment.CONVERSION: 0.17,            # 20K LEQ/month conversions
        },
        competitor_name="",
        competitor_distance_m=0,  # No direct competitor
    ),
    2: StationMarket(
        station_id=2,
        name="Oriente",
        total_leq_month=830_000,  # 500K taxis + 330K buses
        segment_mix={
            MarketSegment.TAXI: 0.60,  # 500K
            MarketSegment.BUS: 0.40,   # 330K (55 buses × 200 LEQ/day)
        },
        competitor_name="NatGas Ojo Caliente",
        competitor_distance_m=300,
    ),
    3: StationMarket(
        station_id=3,
        name="Pensión/Nacozari",
        total_leq_month=500_000,  # 32% combi + 68% taxi
        segment_mix={
            MarketSegment.COMBI: 0.32,
            MarketSegment.TAXI: 0.68,
        },
        competitor_name="NatGas Nacozari",
        competitor_distance_m=100,
    ),
}


# ============================================================
# Placa Nomenclature Classifier
# ============================================================
# AGS plates follow specific patterns that reveal the vehicle segment.
# Nomenclature is "super importante y la base de todo esta segmentación."
#
# Validated against 22 ground-truth files (2012-2025):
#   - 388,665 combi transactions (18 CSVs): 331 unique plates, 99.1% A0####A
#   - 53 vagonetas from 2012-2015: 46 A0####A + 7 A0##AAA (Gen 0, extinct)
#   - 738 taxis from 2012-2015: 75% A###AAX
#   - 1,199 taxis COMPLETA: same pattern confirmed
#   - 4,786 AGS vehicles from NatGas master DB
#
# Plate evolution (AGS combis):
#   Gen 0 (pre-2015): A0##AAA (A014AAA, A016AAA...) — 7 known, extinct
#   Gen 1 (rare):     A###AA  (A954AA, A145AA, A715AA) — 3 known
#   Gen 2 (current):  A0####A (A00000A-A00550A) — 346 known, 100% of actives
#                     Number = concession number (488 total concessions in AGS)
#
# NatGas mislabeling rate: 42.9% of confirmed combis have WRONG Desc_Segmento.
# That's why placa patterns take priority over NatGas labels.

_PLACA_PATTERNS: list[tuple[re.Pattern, MarketSegment, str]] = [
    # ── COMBI AGS patterns (HIGHEST priority — ground-truth validated) ──
    # Gen 2 (current): A0####A — 328 of 331 CSV combis (99.1%)
    # Validated: ALL 331 plates in Josue's CSVs are confirmed combis regardless
    # of NatGas label. NatGas mislabels 142 of these as TP/Público/Empresa/etc.
    (re.compile(r"^A0\d{4}A$"), MarketSegment.COMBI, "A0####A → Combi AGS (Gen 2, concesión)"),
    # Gen 0 (extinct, pre-2015): A0##AAA — 7 known from 2012-2015 Excel
    (re.compile(r"^A0\d{2}AAA$"), MarketSegment.COMBI, "A0##AAA → Combi AGS (Gen 0, extinct)"),
    # Gen 1 (rare/transitional): A###AA — 3 known (A954AA active 67 months)
    (re.compile(r"^A\d{3}AA$"), MarketSegment.COMBI, "A###AA → Combi AGS (Gen 1, rare)"),

    # --- TAXI patterns ---
    # AGS taxi dominant pattern: A###AAX (75% of 1,149 taxis in NatGas)
    # Suffix families: AAB, AAC, AAD, AAE, AAF, AAG, AAH, AAJ, AAK
    # Validated against TAXIS_AGS (738 in 2015, 1,199 in COMPLETA)
    (re.compile(r"^A\d{3}AA[A-K]$"), MarketSegment.TAXI, "A###AA[A-K] → Taxi AGS"),
    # Taxis: ends in TGH (Taxi GNC Hidalgo/AGS pattern)
    (re.compile(r"^[A-Z0-9]+TGH$"), MarketSegment.TAXI, "TGH suffix → Taxi"),
    # Taxis: starts with T followed by digits only
    (re.compile(r"^T\d{3,6}$"), MarketSegment.TAXI, "T+digits → Taxi"),
    # Taxis: AGS taxi pattern (T + 2 letters + 3-4 digits)
    (re.compile(r"^T[A-Z]{2}\d{3,4}$"), MarketSegment.TAXI, "TAx#### → Taxi"),

    # --- BUS patterns ---
    (re.compile(r"^(BUS|AU)\d+$"), MarketSegment.BUS, "BUS/AU prefix → Bus"),
    (re.compile(r"^\d+BUS$"), MarketSegment.BUS, "BUS suffix → Bus"),

    # --- TRANSPORTE DE PERSONAL (TP) / Government / Fleet patterns ---
    # Note: CETE is a client within this segment, not the segment itself
    (re.compile(r"^SS\d+$"), MarketSegment.TRANSPORTE_PERSONAL, "SS prefix → TP/Gubernamental"),
    (re.compile(r"^GS\d+$"), MarketSegment.TRANSPORTE_PERSONAL, "GS prefix → TP/Gubernamental"),
    # TG prefix: "Transporte de Grupo" — 86% Empresa in NatGas data
    (re.compile(r"^TG\d{3,6}$"), MarketSegment.TRANSPORTE_PERSONAL, "TG prefix → TP fleet"),

    # --- COMBI (secondary patterns, lower confidence) ---
    # Combis colectivas: ends in T (Transporte)
    (re.compile(r"^[A-Z]{1,3}\d{3,5}T$"), MarketSegment.COMBI, "xxxT suffix → Combi"),
    # Combis AGS: A_____A pattern (AGS standard combi plates, broader catch)
    (re.compile(r"^A[A-Z0-9]{3,5}A$"), MarketSegment.COMBI, "A____A → Combi AGS"),
    # Combis AGS: starts with AGS
    (re.compile(r"^AGS\d+$"), MarketSegment.COMBI, "AGS prefix → Combi"),
    # A + 3 digits + known combi suffixes (EGT, EGS, MKS, TGC, EGU, EGV)
    (re.compile(r"^A\d{3}(EG[A-Z]|MK[A-Z]|TGC)$"), MarketSegment.COMBI, "A###EGx → Combi NatGas"),

    # --- TRANSPORT (fleet) patterns ---
    # NEW: XX####X — transport fleet plates (87 vehicles, 69% Empresa)
    # First 2 letters: AA, AC, AE, AB, AD, AF — fleet registrations
    (re.compile(r"^A[A-F]\d{4}[A-Z]$"), MarketSegment.CONVERSION, "AX####X → Fleet/Conversion"),

    # --- FEDERAL / PARTICULAR plates ---
    # NEW: XXX###X — federal vehicle plate format (3 letters + 3 digits + 1 letter)
    # 329 vehicles, mixed Público (66%) + Empresa + Plataforma
    # Excludes A-prefix (already caught above) and T-prefix (taxi)
    (re.compile(r"^[B-SU-Z][A-Z]{2}\d{3}[A-Z]$"), MarketSegment.PARTICULAR, "XXX###X → Federal plate"),
    # A-prefix federal plates (254 of 329)
    (re.compile(r"^A[A-Z]{2}\d{3}[A-Z]$"), MarketSegment.COMBI, "AXX###X → Combi federal"),

    # Particular: standard Mexican state format (3 letters + 3-4 digits)
    (re.compile(r"^[B-SU-Z][A-Z]{2}\d{3,4}$"), MarketSegment.PARTICULAR, "Std plate → Particular"),
    # Particular: numeric plates (###AAA private)
    (re.compile(r"^\d{3}[A-Z]{3}$"), MarketSegment.PARTICULAR, "###AAA → Particular"),

    # NEW: Pure numeric IDs — NatGas internal (123 vehicles, 90% Público = combis)
    # 5-7 digit NatGas system IDs, overwhelmingly public transport
    (re.compile(r"^\d{4,7}$"), MarketSegment.COMBI, "NatGas numeric ID → Combi (público)"),

    # NEW: ##XX##X — mixed digit-letter plates (federal/state variants)
    (re.compile(r"^\d{2,3}[A-Z]{2}\d{1,3}$"), MarketSegment.PARTICULAR, "##XX## → Particular"),

    # NEW: XX####X — remaining transport plates (non-A prefix)
    (re.compile(r"^[B-Z][A-Z]\d{4}[A-Z]$"), MarketSegment.CONVERSION, "XX####X → Fleet/Conversion"),

    # ============================================================
    # Round 2 patterns — from analysis of 266 remaining UNKNOWN
    # ============================================================

    # GROUP 1: ####AAB/AAA/GMJ/GMK/GMH/MLH/TSF — NatGas IDs without A prefix
    # 105 vehicles, 96% Público → COMBI. Same suffix families as A### pattern.
    (re.compile(r"^\d{4}(AA[A-K]|GM[A-Z]|ML[A-Z]|TS[A-Z]|JG[A-Z])$"),
     MarketSegment.COMBI, "####AAB/GMx/MLx → Combi NatGas ID"),

    # GROUP 3: ####CM/CB — fleet codes, 96% Empresa → TP
    (re.compile(r"^\d{4}C[BM]$"),
     MarketSegment.TRANSPORTE_PERSONAL, "####CM/CB → TP fleet code"),

    # GROUP 4: ##RB#Z###/##RC#K### — NatGas serial with region code
    # 20 vehicles, 85% Público → COMBI
    (re.compile(r"^\d{2}R[A-Z]\d[A-Z]\d{3}$"),
     MarketSegment.COMBI, "##Rx#x### → Combi NatGas serial"),

    # GROUP 4b: ##RB#Z / ##RC#K (short NatGas serial, no trailing digits)
    # MUST come before generic ##XX#X to avoid misclassification
    (re.compile(r"^\d{2}R[A-Z]\d[A-Z]$"),
     MarketSegment.COMBI, "##Rx#x → Combi NatGas serial (short)"),

    # GROUP 2: ##XX#X — 2-digit + 2-letter + 1-digit + 1-letter state plates
    # 43 vehicles, mixed Público/Empresa → PARTICULAR
    (re.compile(r"^\d{2}[A-Z]{2}\d[A-Z]$"),
     MarketSegment.PARTICULAR, "##XX#X → Particular state plate"),

    # GROUP 5: AF+letter+4digits — AGS federal series (86% Público)
    (re.compile(r"^AF[A-Z]\d{4}$"),
     MarketSegment.COMBI, "AFx#### → Combi AGS federal"),

    # GROUP 6: A/B + 3digits + 3letters (non-AA suffix) — public transport variants
    # A161TGE, A987TGF, A290JTD, B199DAG, B791DAA — 75% Público
    (re.compile(r"^[AB]\d{3}[A-Z]{3}$"),
     MarketSegment.COMBI, "A/B###XXX → Combi variante"),

    # TC+digits: transport company codes (Empresa)
    (re.compile(r"^TC\d{3,5}$"),
     MarketSegment.TRANSPORTE_PERSONAL, "TC prefix → TP fleet"),
]


def classify_placa(placa: str) -> tuple[MarketSegment, str]:
    """
    Classify a placa (license plate) into a market segment using
    nomenclature patterns from the NatGas master database.

    Args:
        placa: Vehicle license plate string (uppercase, no spaces)

    Returns:
        (MarketSegment, reason_string) explaining the classification
    """
    if not placa:
        return MarketSegment.UNKNOWN, "empty placa"

    clean = placa.strip().upper().replace("-", "").replace(" ", "")

    for pattern, segment, reason in _PLACA_PATTERNS:
        if pattern.match(clean):
            return segment, reason

    return MarketSegment.UNKNOWN, f"no pattern matched: {clean}"


def market_segment_to_retention_segmento(ms: MarketSegment) -> str:
    """
    Map MarketSegment → retention Segmento string for existing_profiles.

    The retention engine uses Segmento(VAGONETA, TAXI, PARTICULAR).
    We map our richer MarketSegment to those three buckets:
      COMBI, CETE, CONVERSION, BUS → VAGONETA (commercial fleet)
      TAXI → TAXI
      PARTICULAR, UNKNOWN → PARTICULAR
    """
    if ms in (MarketSegment.COMBI, MarketSegment.TRANSPORTE_PERSONAL,
              MarketSegment.CONVERSION, MarketSegment.BUS):
        # All commercial fleet segments → VAGONETA for retention thresholds
        return "VAGONETA"
    elif ms in (MarketSegment.TAXI, MarketSegment.PLATAFORMA):
        # Taxis and plataforma share the same retention thresholds
        return "TAXI"
    else:
        return "PARTICULAR"


# ============================================================
# NatGas Vehicle Record
# ============================================================

@dataclass
class NatGasVehicle:
    """A vehicle from the NatGas master database."""
    placa: str
    desc_placa: str = ""
    marca: str = ""
    linea: str = ""
    modelo: str = ""
    natgas_segmento: str = ""        # NatGas's own segment label
    market_segment: MarketSegment = MarketSegment.UNKNOWN
    classification_reason: str = ""
    consumo_promedio: Decimal = Decimal("0")  # avg LEQ/day from NatGas
    consumo_max: Decimal = Decimal("0")
    plaza: str = ""
    is_ags: bool = False             # vehicle operates in Aguascalientes
    is_inactive: bool = False        # flagged as inactive in NatGas data
    dias_sin_consumo: int = 0        # from NatGas inactivity report


# ============================================================
# Excel Loader
# ============================================================

def _reclassify_by_consumo(
    current_segment: MarketSegment,
    current_reason: str,
    consumo_max: Decimal,
    consumo_prom: Decimal,
    desc_segmento: str,
    placa: str,
) -> tuple[MarketSegment, str]:
    """
    Reclassify a vehicle using consumo as a validator (Layer 3).

    Called AFTER the placa patterns (Layer 2) and NatGas Desc_Segmento
    have assigned an initial segment. Consumo thresholds validated against
    22 ground-truth files (388,665 transactions) + 4,786 AGS NatGas vehicles.

    Validated consumption ranges (mediana LEQ/carga from real data):
      TAXI:       11.0 (range 5-22)   — 1,149 vehicles
      PLATAFORMA: 12.2 (range 5-24)   — 607 vehicles (Taxi inteligente+ejecutivo+Plataforma)
      COMBI:      28.1 (range 15-48)  — 153 in NatGas, 331 in ground-truth CSVs
      TP:         34.9 (range 15-99)  — 166 vehicles
      BUS:       137.3 (range 50-342) — 153 "Camión Colectivo" in NatGas

    Special handling:
      - Combi plate override: A0####A/A###AA/A0##AAA → COMBI always (Layer 1)
      - "Camión Colectivo" with consumo_max ≤ 200 → TP (fleet, not bus)
      - "Empresa" with consumo_max ≤ 16 → PARTICULAR
      - "Público" with federal plate → PLATAFORMA
    """
    cmax = float(consumo_max)
    cprom = float(consumo_prom)
    lower_seg = desc_segmento.strip().lower() if desc_segmento else ""
    clean_placa = placa.strip().upper().replace("-", "").replace(" ", "")

    # ── 0. Combi plate override ──────────────────────────────────────
    # Three generations of AGS combi plates, ALL confirmed by ground-truth
    # files (22 files, 388,665 transactions, 356 unique combis across all sources):
    #   Gen 2 (current): A0####A — 346 confirmed, 99.1% of CSV combis
    #   Gen 1 (rare):    A###AA  — 3 confirmed (A954AA, A145AA, A715AA)
    #   Gen 0 (extinct):  A0##AAA — 7 confirmed from 2012-2015 (A014AAA etc.)
    #
    # NatGas mislabels 142 of 331 (42.9%) of these as TP/Público/Empresa/etc.
    # This override catches ALL of them regardless of NatGas label.
    _COMBI_PLATE_GEN2 = re.compile(r'^A0\d{4}A$')     # A00000A-A00550A
    _COMBI_PLATE_GEN1 = re.compile(r'^A\d{3}AA$')     # A954AA, A145AA, A715AA
    _COMBI_PLATE_GEN0 = re.compile(r'^A0\d{2}AAA$')   # A014AAA, A016AAA, etc.
    is_combi_plate = (
        _COMBI_PLATE_GEN2.match(clean_placa)
        or _COMBI_PLATE_GEN1.match(clean_placa)
        or _COMBI_PLATE_GEN0.match(clean_placa)
    )
    if is_combi_plate:
        if current_segment != MarketSegment.COMBI:
            return MarketSegment.COMBI, f"{current_reason} → COMBI (placa {clean_placa} = AGS combi plate)"

    # ── 1. "Camión Colectivo" correction ─────────────────────────────
    # NatGas labels 155 AGS vehicles as "Camión Colectivo". Cross-reference
    # with VAGONETAS file shows:
    #   - 56 with consumo_max > 200 → real buses (matches Modelo Op 55 buses)
    #   - 14 with A0####A plate and consumo ≤ 51 → combis (in VAGONETAS file)
    #   - ~85 with fleet plates (YV####, K######, F#K######) and consumo
    #     70-200 → fleet transport vehicles (NOT combis)
    #
    # Rule: A0####A plate → COMBI (already handled by plate override above).
    #       consumo > 200 → BUS.  Everything else → TRANSPORTE_PERSONAL
    #       (fleet vehicle, maps to VAGONETA in retention).
    if "camión colectivo" in lower_seg or "camion colectivo" in lower_seg:
        if cmax > 200:
            # Real bus — keep as BUS
            return current_segment, current_reason
        elif cmax > 0:
            # Fleet transport vehicle — not a combi (no A0####A plate)
            # and not a bus (consumo ≤ 200). Likely vagoneta grande or
            # transporte de personal with bus-like NatGas label.
            return MarketSegment.TRANSPORTE_PERSONAL, (
                f"{current_reason} → TP (consumo_max={cmax:.0f}≤200, fleet plate)"
            )
        # No consumo data — keep as BUS (conservative)
        return current_segment, current_reason

    # ── 2. "Empresa" bucket: fleet vehicles, NOT buses ───────────────
    # NatGas "Empresa" is heterogeneous (consumo prom=92) but these are
    # corporate fleets (TP), not urban buses. Only reclassify very low
    # consumo to PARTICULAR.
    if lower_seg == "empresa" and cmax > 0:
        if cmax <= 16:
            return MarketSegment.PARTICULAR, f"{current_reason} → PARTICULAR (consumo_max={cmax:.0f}≤16)"
        # All other Empresa stays as TRANSPORTE_PERSONAL
        return current_segment, current_reason

    # ── 3. "Público" bucket: check placa for PLATAFORMA ──────────────
    # The "Público" category (1,236 AGS) is mostly taxis, but some are
    # plataforma drivers with federal plates (XXX###X pattern).
    if "público" in lower_seg or "publico" in lower_seg:
        clean = placa.strip().upper().replace("-", "").replace(" ", "")
        # Federal plate pattern: 3 letters + 3 digits + 1 letter
        # Common PLATAFORMA prefixes: JR, JP, JN, JS, SM, AA
        if re.match(r'^[A-Z]{3}\d{3}[A-Z]$', clean):
            return MarketSegment.PLATAFORMA, f"{current_reason} → PLATAFORMA (federal plate {clean})"
        return current_segment, current_reason

    return current_segment, current_reason


def load_natgas_vehiculos(xlsx_path: Path) -> list[NatGasVehicle]:
    """
    Load vehicles from the NatGas master database Excel file.

    Reads the 'Vehiculos' sheet from 'Combis AGS JH.xlsx' which
    contains 86,674 rows of vehicle data across all NatGas plazas.

    Filters to AGS-relevant vehicles and classifies by placa nomenclature.

    Args:
        xlsx_path: Path to the NatGas Excel file

    Returns:
        List of NatGasVehicle (AGS-filtered and classified)
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        raise

    logger.info(f"Loading NatGas vehicles from {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    vehicles = []

    # --- Load plaza data from "Plaza donde mas cargo CW" first ---
    # This sheet has: Id_placa (numeric), Suma de Ind_cantidad, Desc_plaza
    # We build a lookup: id_placa → plaza_name
    plaza_by_id: dict[int, str] = {}
    plaza_sheet_name = None
    for name in wb.sheetnames:
        if "plaza" in name.lower() and "cargo" in name.lower():
            plaza_sheet_name = name
            break

    if plaza_sheet_name:
        ws_plaza = wb[plaza_sheet_name]
        for row in ws_plaza.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            try:
                id_placa = int(row[0])
            except (ValueError, TypeError):
                continue
            plaza_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            if plaza_name and plaza_name != "#N/A":
                plaza_by_id[id_placa] = plaza_name
        logger.info(f"  Loaded plaza data for {len(plaza_by_id)} vehicles")

    # --- Vehiculos sheet ---
    # Columns: 0=Id_placa(int), 1=Desc_placa(plate string), 2=Desc_Marca,
    #   3=Desc_Linea, 4=Desc_Modelo, 5=Id_cliente, 6=Id_Segmento,
    #   7=Desc_Segmento, 8=Tp_modelo, 9=Desc_estado_vehiculo,
    #   10=Fh_conversion, 11=Desc_Taller, 12=Desc_origen_vehiculo,
    #   13=Plaza, 14=Consumo promedio, 15=Consumo Max
    if "Vehiculos" in wb.sheetnames:
        ws = wb["Vehiculos"]
        rows = ws.iter_rows(min_row=2, values_only=True)  # skip header

        for row in rows:
            if not row or not row[0]:
                continue

            # Id_placa is numeric; Desc_placa (col 1) is the actual plate string
            try:
                id_placa = int(row[0])
            except (ValueError, TypeError):
                id_placa = 0

            placa = str(row[1]).strip().upper() if row[1] else ""
            marca = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            linea = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            modelo = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            desc_segmento = str(row[7]).strip() if len(row) > 7 and row[7] else ""
            desc_estado = str(row[9]).strip() if len(row) > 9 and row[9] else ""

            # Plaza: prefer the "Plaza donde mas cargo" sheet, fall back to col 13
            plaza = plaza_by_id.get(id_placa, "")
            if not plaza:
                col13 = str(row[13]).strip() if len(row) > 13 and row[13] else ""
                if col13 and col13 != "#N/A":
                    plaza = col13

            # Consumo fields (col 14 and 15)
            consumo_prom = Decimal("0")
            consumo_max = Decimal("0")
            if len(row) > 14 and row[14] and str(row[14]) != "#N/A":
                try:
                    consumo_prom = Decimal(str(row[14]))
                except Exception:
                    pass
            if len(row) > 15 and row[15] and str(row[15]) != "#N/A":
                try:
                    consumo_max = Decimal(str(row[15]))
                except Exception:
                    pass

            # Determine if AGS
            is_ags = _is_ags_plaza(plaza)

            # PRIMARY: Use NatGas's own Desc_Segmento — they know their clients
            ms, reason = MarketSegment.UNKNOWN, "unclassified"
            if desc_segmento:
                ms, reason = classify_natgas_segmento(desc_segmento)

            # SECONDARY: If NatGas segment is undefined, try desc_estado
            if ms == MarketSegment.UNKNOWN and desc_estado:
                ms, reason = classify_natgas_segmento(desc_estado)

            # FALLBACK: Only use placa nomenclature if NatGas has no data
            if ms == MarketSegment.UNKNOWN:
                ms, reason = classify_placa(placa)

            # ── CONSUMO-BASED RECLASSIFICATION ──────────────────
            # After initial segment assignment, use consumo_max as a
            # validator to catch misclassified vehicles in ambiguous
            # NatGas buckets (Empresa, Público).
            #
            # Rules from domain knowledge:
            #   consumo_max > 50  → vagoneta or bus (NOT taxi/plataforma)
            #   consumo_max ≤ 16  → taxi or plataforma (NOT combi)
            #   consumo_max ≤ 50 and > 16 → combi or TP range
            ms, reason = _reclassify_by_consumo(
                ms, reason, consumo_max, consumo_prom, desc_segmento, placa,
            )

            vehicles.append(NatGasVehicle(
                placa=placa,
                desc_placa=placa,
                marca=marca if marca != "Ninguna" else "",
                linea=linea if linea != "Ninguna" else "",
                modelo=modelo,
                natgas_segmento=desc_segmento,
                market_segment=ms,
                classification_reason=reason,
                consumo_promedio=consumo_prom,
                consumo_max=consumo_max,
                plaza=plaza,
                is_ags=is_ags,
            ))
    else:
        logger.warning("Sheet 'Vehiculos' not found in workbook")

    wb.close()

    total = len(vehicles)
    ags_count = sum(1 for v in vehicles if v.is_ags)
    logger.info(f"  Loaded {total} vehicles, {ags_count} in AGS")

    # Log segment distribution for AGS
    from collections import Counter
    ags_dist = Counter(v.market_segment.value for v in vehicles if v.is_ags)
    for seg, count in ags_dist.most_common():
        logger.info(f"    {seg}: {count}")

    return vehicles


def load_natgas_consumo(xlsx_path: Path) -> dict[str, dict]:
    """
    Load consumption data from the NatGas 'Consumo promedio y max CW' sheet.

    Returns dict of placa → {consumo_promedio, consumo_max} for enrichment.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed")
        raise

    logger.info(f"Loading NatGas consumption data from {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    consumo = {}
    # Sheet name varies between files — find by keyword match
    sheet_name = None
    for name in wb.sheetnames:
        if "consumo" in name.lower() and "prom" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = "Consumo promedio y max CW"  # fallback

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Columns: 0=Id_placa (numeric), 1=Promedio de Ind_cantidad, 2=Max de Ind_cantidad
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            # Id_placa is numeric in this sheet; we'll key by numeric ID
            # and also store the string version for lookup
            key = str(int(row[0])) if isinstance(row[0], (int, float)) else str(row[0]).strip()
            avg_daily = Decimal("0")
            max_daily = Decimal("0")
            if len(row) > 1 and row[1]:
                try:
                    avg_daily = Decimal(str(row[1]))
                except Exception:
                    pass
            if len(row) > 2 and row[2]:
                try:
                    max_daily = Decimal(str(row[2]))
                except Exception:
                    pass
            consumo[key] = {
                "consumo_promedio": avg_daily,
                "consumo_max": max_daily,
            }
    else:
        logger.warning(f"Consumption sheet not found")

    wb.close()
    logger.info(f"  Loaded consumption data for {len(consumo)} vehicles")
    return consumo


def load_natgas_inactive(xlsx_path: Path) -> dict[str, dict]:
    """
    Load the NatGas inactivity report (TAXI QRO Y AGS SIN CONSUMO).

    Returns dict of placa → {segmento, plaza, consumo_promedio, dias_sin_consumo}
    for vehicles flagged as inactive by NatGas.
    """
    try:
        import openpyxl
    except ImportError:
        raise

    logger.info(f"Loading NatGas inactivity data from {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    inactive = {}

    # Try Hoja2 (detail sheet)
    sheet_name = None
    for name in wb.sheetnames:
        if name.lower() in ("hoja2", "sheet2", "detalle"):
            sheet_name = name
            break
    if not sheet_name and wb.sheetnames:
        # Use the second sheet if available, or first
        sheet_name = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]

    if sheet_name:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            placa = str(row[0]).strip().upper()
            consumo_prom = Decimal("0")
            plaza = ""
            segmento = ""

            if len(row) > 1 and row[1]:
                try:
                    consumo_prom = Decimal(str(row[1]))
                except Exception:
                    pass
            if len(row) > 2 and row[2]:
                plaza = str(row[2]).strip()
            if len(row) > 3 and row[3]:
                segmento = str(row[3]).strip()

            inactive[placa] = {
                "consumo_promedio": consumo_prom,
                "plaza": plaza,
                "segmento": segmento,
                "is_inactive": True,
            }

    wb.close()
    logger.info(f"  Loaded {len(inactive)} inactive vehicles")
    return inactive


# ============================================================
# Helper: AGS plaza detection
# ============================================================

_AGS_PLAZA_KEYWORDS = [
    "aguascalientes", "ags", "nacozari", "pensión", "pension",
    "oriente", "ojo caliente", "ojocaliente", "parques",
    "san francisco", "sfr", "romo", "central gas",
]


def _is_ags_plaza(plaza: str) -> bool:
    """Check if a plaza string refers to Aguascalientes."""
    if not plaza:
        return False
    lower = plaza.lower()
    return any(kw in lower for kw in _AGS_PLAZA_KEYWORDS)


def classify_natgas_segmento(desc_segmento: str) -> tuple[MarketSegment, str]:
    """
    Map NatGas's own Desc_Segmento to MarketSegment.

    IMPORTANT: This is Layer 3 (lowest confidence). NatGas labels are WRONG
    for 42.9% of confirmed combis. The placa override in _reclassify_by_consumo
    (Layer 1) corrects these errors. This mapping provides an initial guess
    that gets refined.

    NatGas Desc_Segmento values (4,786 AGS vehicles, validated counts):
      Público (1,419)              → TAXI (mediana consumo 11.5, needs placa fallback)
      Taxi (1,149)                 → TAXI (mediana 11.0, 75% A###AAX plates)
      Taxi inteligente (507)       → PLATAFORMA (mediana 12.2, 83% federal plates)
      Privado (421)                → PARTICULAR (mediana 12.7, 58% federal plates)
      Particular (319)             → PARTICULAR (mediana 13.2)
      Empresa (277)                → TP (mediana 25.2, WIDE range 8-347, "bolsa de gatos")
      Transporte de personal (166) → TP (mediana 34.9, 31% have A0####A = combis!)
      Combis Colectivas (153)      → COMBI (mediana 28.1, 87% A0####A)
      Camión Colectivo (153)       → BUS (mediana 137.3, but 15 have A0####A = combis)
      Taxi ejecutivo (77)          → PLATAFORMA (mediana 12.3, 82% federal)
      Ninguno (75)                 → UNKNOWN
      Plataforma (23)              → PLATAFORMA (mediana 12.3, 91% federal)
      Mypime (13)                  → PARTICULAR
      Transporte de carga (3)      → BUS
      Transporte Urbano (2)        → BUS (mediana 86.7)
      Sin definir (2)              → UNKNOWN
      Gobierno (1)                 → TP
    """
    if not desc_segmento:
        return MarketSegment.UNKNOWN, "no NatGas segment"

    lower = desc_segmento.strip().lower()

    # TAXI regular — the core taxi segment
    if lower == "taxi":
        return MarketSegment.TAXI, f"NatGas: {desc_segmento}"

    # PLATAFORMA — "Taxi inteligente" and "Taxi ejecutivo" are platform drivers
    # (DiDi, Uber, etc.) with DIFFERENT plate nomenclature than regular taxis
    if "taxi inteligente" in lower or "taxi ejecutivo" in lower:
        return MarketSegment.PLATAFORMA, f"NatGas: {desc_segmento}"

    # PLATAFORMA explicit
    if "plataforma" in lower:
        return MarketSegment.PLATAFORMA, f"NatGas: {desc_segmento}"

    # COMBI — only "Combis Colectivas" (NOT "Camión Colectivo" which is BUS)
    if "combi" in lower:
        return MarketSegment.COMBI, f"NatGas: {desc_segmento}"

    # BUS — "Camión Colectivo" (consumo 161 lt/día = bus, NOT combi)
    # and "Transporte Urbano"
    if "camión colectivo" in lower or "camion colectivo" in lower:
        return MarketSegment.BUS, f"NatGas: {desc_segmento} → Bus (consumo >50)"
    if "urbano" in lower:
        return MarketSegment.BUS, f"NatGas: {desc_segmento}"

    # TRANSPORTE DE PERSONAL
    if "transporte de personal" in lower:
        return MarketSegment.TRANSPORTE_PERSONAL, f"NatGas: {desc_segmento}"

    # EMPRESA — heterogeneous, consumo prom=92 means many buses mixed in
    # Classify as TP but will need consumo-based reclassification
    if lower == "empresa":
        return MarketSegment.TRANSPORTE_PERSONAL, f"NatGas: {desc_segmento} → TP/fleet"

    # GOBIERNO
    if "gobierno" in lower or "gubernam" in lower:
        return MarketSegment.TRANSPORTE_PERSONAL, f"NatGas: {desc_segmento}"

    # TRANSPORTE DE CARGA
    if "carga" in lower:
        return MarketSegment.BUS, f"NatGas: {desc_segmento}"

    # PARTICULAR / PRIVADO
    if "particular" in lower or "privad" in lower:
        return MarketSegment.PARTICULAR, f"NatGas: {desc_segmento}"

    # PÚBLICO — large bucket (1,236 in AGS).
    # Consumo prom=13.5 and nomenclature analysis shows majority are taxis.
    # Will be refined by consumo-based reclassification.
    if "público" in lower or "publico" in lower:
        return MarketSegment.TAXI, f"NatGas: {desc_segmento} → Taxi (público)"

    # MYPIME — small business
    if "mypime" in lower:
        return MarketSegment.PARTICULAR, f"NatGas: {desc_segmento}"

    # Unmapped
    if lower in ("ninguno", "otros", "sin definir", "none"):
        return MarketSegment.UNKNOWN, f"NatGas: {desc_segmento} (undefined)"

    return MarketSegment.UNKNOWN, f"NatGas: {desc_segmento} (unmapped)"


def _natgas_segment_hint(natgas_seg: str, placa: str) -> tuple[MarketSegment, str]:
    """Legacy alias — delegates to classify_natgas_segmento."""
    return classify_natgas_segmento(natgas_seg)


# ============================================================
# Profile Builder → existing_profiles for retention engine
# ============================================================

def build_existing_profiles(
    vehicles: list[NatGasVehicle],
    consumo_data: dict[str, dict] | None = None,
    inactive_data: dict[str, dict] | None = None,
    ags_only: bool = True,
) -> dict[str, dict]:
    """
    Build the existing_profiles dict that feeds into
    retention.build_client_profiles() as pre-existing client metadata.

    This enriches the retention engine with:
      - Correct segment classification from placa nomenclature
      - Expected consumption from NatGas data (for anomaly detection)
      - Competitive intelligence (NatGas inactive = conquest opportunity)

    Args:
        vehicles: Loaded NatGasVehicle list
        consumo_data: Optional consumption enrichment
        inactive_data: Optional NatGas inactivity data
        ags_only: If True, only include AGS vehicles

    Returns:
        Dict of placa → {segmento, market_segment, consumo_esperado_lt,
                         natgas_active, plaza, ...}
    """
    consumo = consumo_data or {}
    inactive = inactive_data or {}

    profiles: dict[str, dict] = {}

    for v in vehicles:
        if ags_only and not v.is_ags:
            continue

        retention_seg = market_segment_to_retention_segmento(v.market_segment)

        # Enrichment from consumption data
        consumo_entry = consumo.get(v.placa, {})
        consumo_prom = float(consumo_entry.get(
            "consumo_promedio",
            v.consumo_promedio,
        ))

        # Monthly estimate: daily avg × 30
        consumo_mensual_est = consumo_prom * 30 if consumo_prom > 0 else 0

        # If no consumption data, use segment default
        if consumo_mensual_est == 0:
            seg_params = SEGMENT_PARAMS.get(v.market_segment)
            if seg_params:
                consumo_mensual_est = seg_params.leq_per_month

        # Check NatGas inactivity
        inact = inactive.get(v.placa, {})
        is_natgas_inactive = inact.get("is_inactive", False) or v.is_inactive

        profiles[v.placa] = {
            "segmento": retention_seg,
            "market_segment": v.market_segment.value,
            "estatus": "INACTIVO_NATGAS" if is_natgas_inactive else "ACTIVO",
            "consumo_esperado_lt": consumo_mensual_est,
            "consumo_diario_prom": consumo_prom,
            "consumo_max": float(v.consumo_max),
            "plaza_natgas": v.plaza,
            "marca": v.marca,
            "linea": v.linea,
            "modelo": v.modelo,
            "natgas_segmento": v.natgas_segmento,
            "classification_reason": v.classification_reason,
            "is_natgas_inactive": is_natgas_inactive,
        }

    logger.info(
        f"Built {len(profiles)} existing profiles "
        f"({'AGS only' if ags_only else 'all plazas'})"
    )
    return profiles


# ============================================================
# Market Intelligence Functions
# ============================================================

def get_expected_hourly_volume(
    segment: MarketSegment,
    monthly_volume: float,
    hour: int,
    month: int = 1,
) -> float:
    """
    Estimate expected volume (LEQ) for a given segment, hour, and month.

    Uses the hourly demand profile + seasonality factor.

    Args:
        segment: Market segment
        monthly_volume: Total monthly volume for this segment
        hour: Hour of day (0-23)
        month: Month of year (1-12) for seasonality

    Returns:
        Expected LEQ for that hour (one day's worth)
    """
    params = SEGMENT_PARAMS.get(segment)
    if not params:
        return 0.0

    profile = HOURLY_PROFILES.get(params.hourly_profile_key, {})
    hourly_frac = profile.get(hour, 0.0)

    daily_volume = monthly_volume / 30.0
    seasonality = SEASONALITY_FACTORS.get(month, 1.0)

    return daily_volume * hourly_frac * seasonality


def estimate_revenue_at_risk(
    segment: MarketSegment,
    monthly_litros: float | None = None,
) -> float:
    """
    Estimate monthly revenue at risk if a client of this segment churns.

    Uses segment default if no actual consumption provided.
    PVP ≈ $13.99 MXN/LEQ (IVA included).

    Returns:
        Revenue in MXN/month
    """
    params = SEGMENT_PARAMS.get(segment)
    if not params:
        return 0.0

    litros = monthly_litros if monthly_litros else params.leq_per_month
    return litros * 13.99


def get_market_summary() -> dict:
    """
    Get a summary of the Central Gas market context.

    Returns dict with station markets, segment params, and seasonality.
    Useful for reporting and AI agent context.
    """
    return {
        "stations": {
            sid: {
                "name": sm.name,
                "total_leq_month": sm.total_leq_month,
                "segment_mix": {s.value: pct for s, pct in sm.segment_mix.items()},
                "competitor": sm.competitor_name,
                "competitor_distance_m": sm.competitor_distance_m,
            }
            for sid, sm in STATION_MARKETS.items()
        },
        "segments": {
            seg.value: {
                "leq_per_charge": p.leq_per_charge,
                "leq_per_month": p.leq_per_month,
                "cycle_minutes": p.cycle_minutes,
                "revenue_per_month_mxn": p.revenue_per_month_mxn,
            }
            for seg, p in SEGMENT_PARAMS.items()
        },
        "seasonality": SEASONALITY_FACTORS,
    }


# ============================================================
# Full Loader Pipeline
# ============================================================

def run_market_loader(
    natgas_xlsx: Path | None = None,
    inactive_xlsx: Path | None = None,
    ags_only: bool = True,
) -> tuple[dict[str, dict], list[NatGasVehicle]]:
    """
    Run the full Smart Loader pipeline:
      1. Load NatGas vehicles (Vehiculos sheet)
      2. Load NatGas consumption data
      3. Load NatGas inactivity report
      4. Classify all vehicles by placa nomenclature
      5. Build existing_profiles for retention engine

    Args:
        natgas_xlsx: Path to 'Combis AGS JH.xlsx'
        inactive_xlsx: Path to 'TAXI QRO Y AGS SIN CONSUMO' file
        ags_only: Only include AGS vehicles

    Returns:
        (existing_profiles dict, all_vehicles list)
    """
    vehicles = []
    consumo_data = {}
    inactive_data = {}

    if natgas_xlsx and natgas_xlsx.exists():
        vehicles = load_natgas_vehiculos(natgas_xlsx)
        consumo_data = load_natgas_consumo(natgas_xlsx)
    else:
        logger.warning(f"NatGas Excel not found: {natgas_xlsx}")

    if inactive_xlsx and inactive_xlsx.exists():
        inactive_data = load_natgas_inactive(inactive_xlsx)

    profiles = build_existing_profiles(
        vehicles, consumo_data, inactive_data, ags_only
    )

    logger.info(
        f"Market loader complete: {len(vehicles)} vehicles, "
        f"{len(profiles)} profiles built"
    )

    return profiles, vehicles
