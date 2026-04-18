#!/usr/bin/env python3
"""
Import clients from BASE_CLIENTES_GNC_AGS_COMPLETA.xlsx → Neon clients table.

Usage:
    # Local (needs DATABASE_URL env):
    python3 scripts/import_clients.py /path/to/BASE_CLIENTES_GNC_AGS_COMPLETA.xlsx

    # Via Fly SSH (file must be in image or uploaded):
    fly ssh console --app central-gas-agent --command \
        "python3 /app/scripts/import_clients.py /app/data/BASE_CLIENTES_GNC_AGS_COMPLETA.xlsx"

Idempotent: uses ON CONFLICT (placa) DO UPDATE to upsert.
"""

from __future__ import annotations

import os
import sys
import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(name)s] %(levelname)s: %(message)s")
logger = logging.getLogger("import_clients")


def safe_str(v, max_len: int = 200) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() in ("none", "ninguna", "ninguno", "nan"):
        return None
    return s[:max_len]


def safe_decimal(v) -> Decimal | None:
    if v is None:
        return None
    try:
        return Decimal(str(v)).quantize(Decimal("0.1"))
    except (InvalidOperation, ValueError):
        return None


def safe_date(v) -> datetime | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    try:
        return datetime.fromisoformat(str(v).strip())
    except (ValueError, TypeError):
        return None


# Segmento mapping from GasData labels → our enum
SEGMENTO_MAP = {
    "Taxi": "TAXI",
    "Taxi inteligente": "TAXI",
    "Público": "VAGONETA",
    "Combis Colectivas": "VAGONETA",
    "Camión Colectivo": "VAGONETA",
    "Privado": "PARTICULAR",
    "Particular": "PARTICULAR",
    "Sin definir": "VAGONETA",
    "Ninguno": "VAGONETA",
}


def parse_xlsx(path: str) -> list[dict]:
    """Parse BASE_CLIENTES_GNC xlsx → list of client dicts."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["TAXIS_AGS"]

    headers = [c.value for c in ws[1]]
    clients = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        placa = safe_str(row[0], 20)
        if not placa:
            continue

        segmento_raw = safe_str(row[13]) or ""
        segmento = SEGMENTO_MAP.get(segmento_raw, "VAGONETA")

        clients.append({
            "placa": placa.upper(),
            "nombre": safe_str(row[1]),
            "telefono": safe_str(row[2], 20),
            "whatsapp": safe_str(row[2], 20),  # same as telefono for now
            "email": safe_str(row[3]),
            "rfc": safe_str(row[4], 13),
            "modelo_vehiculo": safe_str(row[8], 20),
            "fecha_conversion": safe_date(row[10]),
            "segmento": segmento,
            "consumo_prom_lt": safe_decimal(row[12]),
            "eds_principal": None,  # will be derived from transactions later
            "notas": f"Imported from BASE_CLIENTES_GNC. Segmento GasData: {segmento_raw}",
        })

    wb.close()
    logger.info(f"Parsed {len(clients)} clients from {path}")
    return clients


def upsert_clients(clients: list[dict]) -> dict:
    """Upsert clients into Neon. Returns stats."""
    import psycopg2
    from psycopg2.extras import execute_values

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        raise RuntimeError("DATABASE_URL not set")

    conn = psycopg2.connect(db_url)
    inserted = 0
    updated = 0
    errors = []

    try:
        with conn.cursor() as cur:
            sql = """
                INSERT INTO clients (
                    placa, nombre, telefono, whatsapp, email, rfc,
                    modelo_vehiculo, fecha_conversion, segmento,
                    consumo_prom_lt, eds_principal, notas, estatus,
                    updated_at
                ) VALUES %s
                ON CONFLICT (placa) DO UPDATE SET
                    nombre = COALESCE(EXCLUDED.nombre, clients.nombre),
                    telefono = COALESCE(EXCLUDED.telefono, clients.telefono),
                    whatsapp = COALESCE(EXCLUDED.whatsapp, clients.whatsapp),
                    email = COALESCE(EXCLUDED.email, clients.email),
                    rfc = COALESCE(EXCLUDED.rfc, clients.rfc),
                    modelo_vehiculo = COALESCE(EXCLUDED.modelo_vehiculo, clients.modelo_vehiculo),
                    fecha_conversion = COALESCE(EXCLUDED.fecha_conversion, clients.fecha_conversion),
                    segmento = EXCLUDED.segmento,
                    consumo_prom_lt = COALESCE(EXCLUDED.consumo_prom_lt, clients.consumo_prom_lt),
                    notas = EXCLUDED.notas,
                    updated_at = NOW()
            """

            rows = []
            for c in clients:
                rows.append((
                    c["placa"], c["nombre"], c["telefono"], c["whatsapp"],
                    c["email"], c["rfc"], c["modelo_vehiculo"],
                    c["fecha_conversion"], c["segmento"], c["consumo_prom_lt"],
                    c["eds_principal"], c["notas"], "ACTIVO", datetime.utcnow(),
                ))

            # Use page_size to batch
            execute_values(cur, sql, rows, page_size=200)

            # Check how many were inserted vs updated
            total_affected = cur.rowcount
            conn.commit()

            # Count actual clients
            cur.execute("SELECT COUNT(*) FROM clients")
            total_in_db = cur.fetchone()[0]

            logger.info(f"Upsert complete: {total_affected} rows affected, {total_in_db} total clients in DB")
            return {
                "affected": total_affected,
                "total_in_db": total_in_db,
                "input_count": len(clients),
                "errors": errors,
            }
    except Exception as e:
        conn.rollback()
        logger.exception(f"Upsert failed: {e}")
        raise
    finally:
        conn.close()


def main():
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <path_to_BASE_CLIENTES_GNC.xlsx>")
        sys.exit(1)

    path = sys.argv[1]
    if not os.path.exists(path):
        logger.error(f"File not found: {path}")
        sys.exit(1)

    clients = parse_xlsx(path)
    if not clients:
        logger.warning("No clients parsed — nothing to import")
        sys.exit(0)

    result = upsert_clients(clients)
    print(f"\n{'='*50}")
    print(f"Import complete:")
    print(f"  Input:     {result['input_count']} clients")
    print(f"  Affected:  {result['affected']} rows (insert or update)")
    print(f"  Total DB:  {result['total_in_db']} clients")
    if result['errors']:
        print(f"  Errors:    {len(result['errors'])}")
        for e in result['errors'][:5]:
            print(f"    - {e}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
